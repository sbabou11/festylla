"""
Générateur de rapport Excel YllaCash
Usage: python3 generate_rapport.py data.json output.xlsx

data.json contient:
{
  "event": { "nom": "...", "couleur": "#1a6b7a" },
  "spectateurs": [...],
  "transactions": [...],
  "reservations": [...],
  "menu": [...],
  "staff": [...],
  "audit": [...]
}
"""
import json, sys, math
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XlImage

# ── Helpers couleur ──────────────────────────────────────────────────
def hex_to_argb(h):
    h = h.lstrip('#')
    return 'FF' + h.upper()

def darken(h, factor=0.3):
    h = h.lstrip('#')
    r,g,b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
    r = int(r * (1 - factor)); g = int(g * (1 - factor)); b = int(b * (1 - factor))
    return f'{r:02X}{g:02X}{b:02X}'

def lighten(h, factor=0.85):
    h = h.lstrip('#')
    r,g,b = int(h[0:2],16), int(h[2:4],16), int(h[4:6],16)
    r = min(255, int(r + (255 - r) * factor))
    g = min(255, int(g + (255 - g) * factor))
    b = min(255, int(b + (255 - b) * factor))
    return f'{r:02X}{g:02X}{b:02X}'

def euro(centimes):
    return round((centimes or 0) / 100, 2)

def fmt_euro(centimes):
    return f"{euro(centimes):.2f} €"

def now_str():
    return datetime.now().strftime('%d/%m/%Y %H:%M')

# ── Styles ───────────────────────────────────────────────────────────
def make_styles(brand):
    brand_argb   = hex_to_argb(brand)
    brand_light  = hex_to_argb(lighten(brand, 0.88))
    brand_dark   = hex_to_argb(darken(brand, 0.15))
    white        = 'FFFFFFFF'
    gray_bg      = 'FFF8F9FA'
    gray_border  = 'FFE2E8F0'
    text_dark    = 'FF0F172A'
    text_muted   = 'FF64748B'

    thin = Side(style='thin', color=gray_border)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    return {
        'brand': brand_argb, 'brand_light': brand_light, 'brand_dark': brand_dark,
        'white': white, 'gray_bg': gray_bg, 'border': border, 'text_dark': text_dark,
        'text_muted': text_muted,

        # Titre principal (header feuille)
        'h1': {
            'font': Font(name='Arial', bold=True, size=16, color=white),
            'fill': PatternFill('solid', fgColor=brand_argb),
            'alignment': Alignment(horizontal='left', vertical='center'),
        },
        # Sous-titre section
        'h2': {
            'font': Font(name='Arial', bold=True, size=11, color=hex_to_argb(darken(brand, 0.2))),
            'fill': PatternFill('solid', fgColor=brand_light),
            'alignment': Alignment(horizontal='left', vertical='center'),
        },
        # En-tête tableau
        'th': {
            'font': Font(name='Arial', bold=True, size=10, color=white),
            'fill': PatternFill('solid', fgColor=brand_argb),
            'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'border': border,
        },
        # Ligne normale
        'td': {
            'font': Font(name='Arial', size=10, color=text_dark),
            'fill': PatternFill('solid', fgColor=white),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': border,
        },
        # Ligne alternée
        'td_alt': {
            'font': Font(name='Arial', size=10, color=text_dark),
            'fill': PatternFill('solid', fgColor=gray_bg),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': border,
        },
        # KPI valeur
        'kpi_val': {
            'font': Font(name='Arial', bold=True, size=18, color=hex_to_argb(darken(brand, 0.15))),
            'fill': PatternFill('solid', fgColor=brand_light),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': border,
        },
        # KPI label
        'kpi_label': {
            'font': Font(name='Arial', bold=True, size=9, color=text_muted),
            'fill': PatternFill('solid', fgColor=brand_light),
            'alignment': Alignment(horizontal='center', vertical='center'),
        },
        # Total / important
        'total': {
            'font': Font(name='Arial', bold=True, size=11, color=white),
            'fill': PatternFill('solid', fgColor=brand_dark),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': border,
        },
        # Montant positif
        'pos': {
            'font': Font(name='Arial', bold=True, size=10, color='FF065F46'),
            'fill': PatternFill('solid', fgColor='FFD1FAE5'),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': border,
        },
        # Montant négatif
        'neg': {
            'font': Font(name='Arial', bold=True, size=10, color='FF991B1B'),
            'fill': PatternFill('solid', fgColor='FFFEE2E2'),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': border,
        },
        # Numérique centré
        'num': {
            'font': Font(name='Arial', size=10, color=text_dark),
            'alignment': Alignment(horizontal='right', vertical='center'),
            'border': border,
        },
    }

def apply(cell, style_dict):
    for k, v in style_dict.items():
        setattr(cell, k, v)

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

def col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ── Feuille 0 : Page de garde ────────────────────────────────────────
def sheet_garde(wb, data, S):
    ws = wb.active
    ws.title = '📊 Tableau de bord'
    ws.sheet_view.showGridLines = False

    brand = data['event'].get('couleur', '#1a6b7a')
    brand_argb = hex_to_argb(brand)
    brand_light = hex_to_argb(lighten(brand, 0.9))
    white = 'FFFFFFFF'
    text_dark = 'FF0F172A'
    text_muted = 'FF64748B'

    # Colonnes
    for c, w in [('A',2),('B',22),('C',22),('D',22),('E',22),('F',22),('G',4)]:
        col_width(ws, c, w)

    # ── HEADER PRINCIPAL ──
    ws.merge_cells('B1:F3')
    h = ws['B1']
    h.value = f"YllaCash — {data['event'].get('nom','Événement').upper()}"
    h.font = Font(name='Arial', bold=True, size=20, color=white)
    h.fill = PatternFill('solid', fgColor=brand_argb)
    h.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    set_row_height(ws, 1, 20); set_row_height(ws, 2, 20); set_row_height(ws, 3, 20)

    ws.merge_cells('B4:F4')
    sub = ws['B4']
    sub.value = f"Rapport financier complet — Généré le {now_str()}"
    sub.font = Font(name='Arial', size=10, color=hex_to_argb(darken(brand, 0.1)))
    sub.fill = PatternFill('solid', fgColor=hex_to_argb(lighten(brand, 0.8)))
    sub.alignment = Alignment(horizontal='left', vertical='center', indent=2)
    set_row_height(ws, 4, 20)
    set_row_height(ws, 5, 12)

    # ── KPIs ──
    txs = data.get('transactions', [])
    specs = data.get('spectateurs', [])
    resas = data.get('reservations', [])

    total_credits = sum(t.get('montant',0) for t in txs if t.get('type') == 'credit')
    total_ventes  = sum(t.get('montant',0) for t in txs if t.get('type') in ('debit','retrait','benev-retrait'))
    total_soldes  = sum(s.get('solde',0) for s in specs)
    nb_specs      = len(specs)
    nb_resas      = len(resas)
    collected     = len([r for r in resas if r.get('status') == 'collected'])
    taux          = round(collected / nb_resas * 100) if nb_resas else 0
    benev_resas   = [r for r in resas if r.get('isBenev') and r.get('status') == 'collected']
    cout_benev    = sum(
        sum((i.get('prix',0) * i.get('qty',1)) for i in r.get('items',[]))
        for r in benev_resas
    )
    ca_net = total_ventes - cout_benev

    kpis = [
        ('CA Brut encaissé', fmt_euro(total_ventes), '💰'),
        ('Total rechargé',   fmt_euro(total_credits), '💳'),
        ('Soldes restants',  fmt_euro(total_soldes), '🏦'),
        ('Spectateurs',      str(nb_specs), '👥'),
        ('Taux retrait résa',f'{taux}%', '📋'),
        ('CA Net événement', fmt_euro(ca_net), '📈'),
    ]

    row = 6
    set_row_height(ws, row, 15)
    row += 1

    for i, (label, val, icon) in enumerate(kpis):
        col = chr(ord('B') + i)
        if i >= 3:
            col = chr(ord('B') + i - 3)
            if i == 3:
                set_row_height(ws, row + 2, 14)
                row_offset = row + 2 + 1
            row_use = row_offset if i >= 3 else row
        else:
            row_use = row

        # Label
        ws.merge_cells(f'{col}{row_use}:{col}{row_use}')
        lbl = ws[f'{col}{row_use}']
        lbl.value = f'{icon} {label}'
        lbl.font = Font(name='Arial', bold=True, size=9, color=text_muted)
        lbl.fill = PatternFill('solid', fgColor=brand_light)
        lbl.alignment = Alignment(horizontal='center', vertical='center')
        set_row_height(ws, row_use, 16)

        # Valeur
        ws.merge_cells(f'{col}{row_use+1}:{col}{row_use+1}')
        v = ws[f'{col}{row_use+1}']
        v.value = val
        v.font = Font(name='Arial', bold=True, size=16, color=hex_to_argb(darken(brand, 0.1)))
        v.fill = PatternFill('solid', fgColor=brand_light)
        v.alignment = Alignment(horizontal='center', vertical='center')
        set_row_height(ws, row_use+1, 32)

    # ── Tableau Répartition types tx ──
    row_table = 18
    set_row_height(ws, row_table - 1, 14)
    ws.merge_cells(f'B{row_table}:F{row_table}')
    th = ws[f'B{row_table}']
    th.value = '📊  Répartition des transactions par type'
    th.font = Font(name='Arial', bold=True, size=11, color=white)
    th.fill = PatternFill('solid', fgColor=brand_argb)
    th.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    set_row_height(ws, row_table, 22)

    headers = ['Type', 'Nb transactions', 'Montant total (€)', 'Moyenne (€)']
    types = [
        ('💳 Crédit',            'credit'),
        ('🛒 Encaissement',      'debit'),
        ('📦 Retrait résa',      'retrait'),
        ('🎁 Retrait bénévole',  'benev-retrait'),
        ('📋 Réservation',       'reservation'),
        ('❌ Annulation',        'annulation'),
    ]

    for j, h_txt in enumerate(headers):
        col = chr(ord('B') + j)
        cell = ws[f'{col}{row_table+1}']
        cell.value = h_txt
        apply(cell, S['th'])
        set_row_height(ws, row_table+1, 20)

    for k, (label, typ) in enumerate(types):
        txs_t = [t for t in txs if t.get('type') == typ]
        nb = len(txs_t)
        mt = sum(t.get('montant',0) for t in txs_t)
        avg = mt / nb if nb else 0
        row_k = row_table + 2 + k
        vals = [label, nb, f'{euro(mt):.2f}', f'{euro(avg):.2f}']
        for j, val in enumerate(vals):
            col = chr(ord('B') + j)
            cell = ws[f'{col}{row_k}']
            cell.value = val
            st = S['td_alt'] if k % 2 else S['td']
            apply(cell, st)
            if j > 0:
                cell.alignment = Alignment(horizontal='right', vertical='center')
        set_row_height(ws, row_k, 18)

    # ── Graphique Barres CA par type ──
    chart_data_row_start = row_table + 2
    chart_data_row_end   = row_table + 2 + len(types) - 1

    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Volume financier par type de transaction'
    chart.style = 10
    chart.y_axis.title = 'Montant (€)'
    chart.x_axis.title = 'Type'
    chart.shape = 4

    data_ref = Reference(ws, min_col=4, max_col=4,
                         min_row=chart_data_row_start, max_row=chart_data_row_end)
    cats_ref = Reference(ws, min_col=2, max_col=2,
                         min_row=chart_data_row_start, max_row=chart_data_row_end)
    chart.add_data(data_ref)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = brand.lstrip('#')
    chart.width  = 18
    chart.height = 12
    ws.add_chart(chart, 'B27')

    # ── Graphique Camembert Spectateurs vs Bénévoles ──
    pie = PieChart()
    pie.title = 'Réservations : Spectateurs vs Bénévoles'
    pie.style = 10

    nb_spec_resas  = len([r for r in resas if not r.get('isBenev')])
    nb_benev_resas = len([r for r in resas if r.get('isBenev')])

    pie_data_start = row_table + 10
    ws[f'B{pie_data_start}'] = 'Spectateurs'; ws[f'C{pie_data_start}'] = nb_spec_resas
    ws[f'B{pie_data_start+1}'] = 'Bénévoles'; ws[f'C{pie_data_start+1}'] = nb_benev_resas
    for r in [pie_data_start, pie_data_start+1]:
        ws.row_dimensions[r].hidden = True  # Données cachées

    pie_data = Reference(ws, min_col=3, max_col=3,
                         min_row=pie_data_start, max_row=pie_data_start+1)
    pie_cats = Reference(ws, min_col=2, max_col=2,
                         min_row=pie_data_start, max_row=pie_data_start+1)
    pie.add_data(pie_data)
    pie.set_categories(pie_cats)
    pie.width = 14; pie.height = 12
    ws.add_chart(pie, 'D27')

    return ws

# ── Feuille Transactions ─────────────────────────────────────────────
def sheet_transactions(wb, data, S):
    ws = wb.create_sheet('💳 Transactions')
    ws.sheet_view.showGridLines = False

    headers = ['Date','Heure','Type','Spectateur / Bénévole','Libellé','Montant (€)','Staff','Code résa']
    widths  = [14, 10, 18, 22, 40, 14, 18, 14]

    for c, w in zip([get_column_letter(i+1) for i in range(len(widths))], widths):
        col_width(ws, c, w)

    # Titre
    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    t = ws['A1']
    t.value = '💳  Détail complet des transactions'
    apply(t, S['h1'])
    set_row_height(ws, 1, 28)

    ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    ws['A2'].value = f"Généré le {now_str()}"
    ws['A2'].font = Font(name='Arial', size=9, color=S['text_muted'])
    set_row_height(ws, 2, 16)

    # En-têtes
    for j, h_txt in enumerate(headers):
        cell = ws.cell(row=4, column=j+1, value=h_txt)
        apply(cell, S['th'])
    set_row_height(ws, 4, 22)

    txs = sorted(data.get('transactions', []),
                 key=lambda t: t.get('timestamp',''), reverse=True)

    type_labels = {
        'credit':'💳 Crédit','debit':'🛒 Encaissement','retrait':'📦 Retrait',
        'benev-retrait':'🎁 Retrait bénévole','reservation':'📋 Réservation',
        'annulation':'❌ Annulation','benev-reservation':'📋 Résa bénévole',
        'benev-annulation':'❌ Annul. bénévole',
    }

    for i, t in enumerate(txs):
        row = i + 5
        ts = t.get('timestamp','')
        try:
            dt = datetime.fromisoformat(ts)
            date_str = dt.strftime('%d/%m/%Y')
            time_str = dt.strftime('%H:%M')
        except:
            date_str = t.get('date','—')
            time_str = t.get('heure','—')

        mt = t.get('montant', 0)
        who = t.get('benevoleNom') or t.get('specNom') or '—'
        row_vals = [
            date_str, time_str,
            type_labels.get(t.get('type',''), t.get('type','—')),
            who, t.get('label','—'),
            euro(mt), t.get('staff','—'), t.get('resaCode','—'),
        ]
        style = S['td_alt'] if i % 2 else S['td']
        for j, val in enumerate(row_vals):
            cell = ws.cell(row=row, column=j+1, value=val)
            apply(cell, style)
            if j == 5:  # Montant
                cell.number_format = '#,##0.00 "€"'
                if t.get('type') == 'credit':
                    apply(cell, S['pos'])
                elif t.get('type') in ('debit','retrait','benev-retrait'):
                    apply(cell, S['neg'])
                cell.alignment = Alignment(horizontal='right', vertical='center')
        set_row_height(ws, row, 18)

    # Total
    total_row = len(txs) + 5
    ws.cell(row=total_row, column=5, value='TOTAL TRANSACTIONS').font = Font(bold=True, name='Arial', size=10, color='FFFFFFFF')
    ws.cell(row=total_row, column=5).fill = PatternFill('solid', fgColor=S['brand'])
    ws.cell(row=total_row, column=6, value=f'=SUM(F5:F{total_row-1})')
    apply(ws.cell(row=total_row, column=6), S['total'])
    ws.cell(row=total_row, column=6).number_format = '#,##0.00 "€"'
    set_row_height(ws, total_row, 22)

    ws.freeze_panes = 'A5'
    return ws

# ── Feuille Spectateurs ──────────────────────────────────────────────
def sheet_spectateurs(wb, data, S):
    ws = wb.create_sheet('👥 Spectateurs')
    ws.sheet_view.showGridLines = False

    headers = ['ID QR','Nom','Solde actuel (€)','Nb transactions','Total rechargé (€)','Total dépensé (€)','Dernière opération']
    widths  = [24, 24, 16, 14, 18, 18, 20]
    for c, w in zip([get_column_letter(i+1) for i in range(len(widths))], widths):
        col_width(ws, c, w)

    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    t = ws['A1']; t.value = '👥  Registre des spectateurs'; apply(t, S['h1'])
    set_row_height(ws, 1, 28)

    for j, h in enumerate(headers):
        cell = ws.cell(row=3, column=j+1, value=h); apply(cell, S['th'])
    set_row_height(ws, 3, 22)

    txs = data.get('transactions', [])
    specs = sorted(data.get('spectateurs', []), key=lambda s: (s.get('solde',0)), reverse=True)

    for i, s in enumerate(specs):
        row = i + 4
        my_tx = [t for t in txs if t.get('specId') == s.get('id')]
        credits = sum(t.get('montant',0) for t in my_tx if t.get('type') == 'credit')
        debits  = sum(t.get('montant',0) for t in my_tx if t.get('type') in ('debit','retrait'))
        last_ts = max((t.get('timestamp','') for t in my_tx), default='')
        try:
            last = datetime.fromisoformat(last_ts).strftime('%d/%m/%Y %H:%M') if last_ts else '—'
        except: last = '—'

        vals = [s.get('id',''), s.get('nom',''), euro(s.get('solde',0)),
                len(my_tx), euro(credits), euro(debits), last]
        style = S['td_alt'] if i % 2 else S['td']
        for j, val in enumerate(vals):
            cell = ws.cell(row=row, column=j+1, value=val); apply(cell, style)
            if j in (2, 4, 5):
                cell.number_format = '#,##0.00 "€"'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            if j == 2 and isinstance(val, float):
                if val < 0: apply(cell, S['neg'])
                elif val > 0: apply(cell, S['pos'])
        set_row_height(ws, row, 18)

    ws.freeze_panes = 'A4'
    return ws

# ── Feuille Réservations ─────────────────────────────────────────────
def sheet_reservations(wb, data, S):
    ws = wb.create_sheet('📋 Réservations')
    ws.sheet_view.showGridLines = False

    headers = ['Code','Bénéficiaire','Type','Articles','Total (€)','Statut','Date','Pris en charge par']
    widths  = [16, 24, 12, 40, 12, 18, 16, 20]
    for c, w in zip([get_column_letter(i+1) for i in range(len(widths))], widths):
        col_width(ws, c, w)

    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    t = ws['A1']; t.value = '📋  Journal des réservations'; apply(t, S['h1'])
    set_row_height(ws, 1, 28)

    for j, h in enumerate(headers):
        cell = ws.cell(row=3, column=j+1, value=h); apply(cell, S['th'])
    set_row_height(ws, 3, 22)

    STATUS = {
        'pending':'⏳ En revue','processing':'👨‍🍳 En préparation',
        'ready':'✅ Prête','collected':'📦 Retirée','cancelled':'❌ Annulée',
    }
    STATUS_COLORS = {
        'pending':'FFFEF3C7','processing':'FFEDE9FE',
        'ready':'FFD1FAE5','collected':'FFF1F5F9','cancelled':'FFFEE2E2',
    }

    resas = sorted(data.get('reservations', []),
                   key=lambda r: r.get('date',''), reverse=True)

    for i, r in enumerate(resas):
        row = i + 4
        status = r.get('status','')
        items = ', '.join(f"{it.get('nom','')} ×{it.get('qty',1)}" for it in r.get('items',[]))
        beneficiaire = r.get('benevoleNom') or r.get('specNom') or '—'
        typ = '🙋 Bénévole' if r.get('isBenev') else '👥 Spectateur'
        mt = r.get('total', 0)
        if not mt or mt != mt:
            mt = sum((it.get('prix',0) * it.get('qty',1)) for it in r.get('items',[]))

        vals = [r.get('code',''), beneficiaire, typ, items,
                euro(mt), STATUS.get(status, status), r.get('date',''), r.get('collectedBy','—')]

        style = S['td_alt'] if i % 2 else S['td']
        for j, val in enumerate(vals):
            cell = ws.cell(row=row, column=j+1, value=val); apply(cell, style)
            if j == 4:
                cell.number_format = '#,##0.00 "€"'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            if j == 5 and status in STATUS_COLORS:
                cell.fill = PatternFill('solid', fgColor=STATUS_COLORS[status])
                cell.alignment = Alignment(horizontal='center', vertical='center')
        set_row_height(ws, row, 18)

    ws.freeze_panes = 'A4'
    return ws

# ── Feuille Menu & Stocks ────────────────────────────────────────────
def sheet_menu(wb, data, S):
    ws = wb.create_sheet('🍽️ Menu & Stocks')
    ws.sheet_view.showGridLines = False

    headers = ['Article','Catégorie','Type conso','Prix (€)','Stock restant','Unités vendues','CA généré (€)']
    widths  = [28, 18, 16, 12, 14, 16, 16]
    for c, w in zip([get_column_letter(i+1) for i in range(len(widths))], widths):
        col_width(ws, c, w)

    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    t = ws['A1']; t.value = '🍽️  Menu et performances des articles'; apply(t, S['h1'])
    set_row_height(ws, 1, 28)

    for j, h in enumerate(headers):
        cell = ws.cell(row=3, column=j+1, value=h); apply(cell, S['th'])
    set_row_height(ws, 3, 22)

    txs = data.get('transactions', [])
    menu_items = sorted(data.get('menu', []), key=lambda m: m.get('cat',''))

    for i, m in enumerate(menu_items):
        row = i + 4
        ventes = [it for t in txs
                  if t.get('type') in ('debit','retrait','benev-retrait')
                  for it in t.get('items',[]) if it.get('nom') == m.get('nom')]
        unites = sum(it.get('qty',1) for it in ventes)
        ca     = sum(it.get('total', it.get('prixUnit',0) * it.get('qty',1)) for it in ventes)

        type_map = {'repas':'🍽️ Repas','boisson':'☕ Boisson','eau':'💧 Eau','':'—'}
        vals = [m.get('nom',''), m.get('cat',''), type_map.get(m.get('typeConsommation',''),'—'),
                euro(m.get('prix',0)), m.get('stock',0), unites, euro(ca)]

        style = S['td_alt'] if i % 2 else S['td']
        for j, val in enumerate(vals):
            cell = ws.cell(row=row, column=j+1, value=val); apply(cell, style)
            if j in (3, 6):
                cell.number_format = '#,##0.00 "€"'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            if j in (4, 5):
                cell.alignment = Alignment(horizontal='center', vertical='center')
        set_row_height(ws, row, 18)

    # Graphique Top ventes
    if menu_items:
        n = min(len(menu_items), 10)
        top = sorted(menu_items, key=lambda m: sum(
            sum(it.get('qty',1) for it in t.get('items',[]) if it.get('nom')==m.get('nom'))
            for t in txs if t.get('type') in ('debit','retrait','benev-retrait')
        ), reverse=True)[:n]

        chart_row = len(menu_items) + 6
        ws.cell(row=chart_row, column=1).value = 'Article'
        ws.cell(row=chart_row, column=2).value = 'Unités vendues'
        for k, m in enumerate(top):
            ventes = [it for t in txs if t.get('type') in ('debit','retrait','benev-retrait')
                      for it in t.get('items',[]) if it.get('nom') == m.get('nom')]
            ws.cell(row=chart_row+1+k, column=1).value = m.get('nom','')
            ws.cell(row=chart_row+1+k, column=2).value = sum(it.get('qty',1) for it in ventes)
            ws.row_dimensions[chart_row+1+k].hidden = True
        ws.row_dimensions[chart_row].hidden = True

        chart = BarChart()
        chart.type = 'bar'
        chart.title = f'Top {n} articles — Unités vendues'
        chart.style = 10
        chart.y_axis.title = 'Unités'
        chart.x_axis.title = 'Article'
        chart.width = 22; chart.height = 14

        data_ref = Reference(ws, min_col=2, max_col=2, min_row=chart_row, max_row=chart_row+n)
        cats_ref = Reference(ws, min_col=1, max_col=1, min_row=chart_row+1, max_row=chart_row+n)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        if chart.series:
            chart.series[0].graphicalProperties.solidFill = S['brand'].lstrip('FF')
        ws.add_chart(chart, 'A5')

    ws.freeze_panes = 'A4'
    return ws

# ── Feuille Audit ────────────────────────────────────────────────────
def sheet_audit(wb, data, S):
    ws = wb.create_sheet('📋 Journal d\'audit')
    ws.sheet_view.showGridLines = False

    headers = ['Date','Heure','Action','Type utilisateur','Libellé','Spectateur/Bénévole','Staff','Code résa','Montant (€)']
    widths  = [14,9,22,16,44,26,20,14,12]
    for c, w in zip([get_column_letter(i+1) for i in range(len(widths))], widths):
        col_width(ws, c, w)

    ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
    t = ws['A1']; t.value = "📋  Journal d'audit complet"; apply(t, S['h1'])
    set_row_height(ws, 1, 28)

    for j, h in enumerate(headers):
        cell = ws.cell(row=3, column=j+1, value=h); apply(cell, S['th'])
    set_row_height(ws, 3, 22)

    logs = sorted(data.get('audit', []),
                  key=lambda l: l.get('timestamp',''), reverse=True)

    USER_ICON = {'staff':'👤','admin':'🛡️','benevole':'🙋','spectateur':'👥'}
    for i, l in enumerate(logs):
        row = i + 4
        ts = l.get('timestamp','')
        try:
            dt = datetime.fromisoformat(ts)
            date_str = dt.strftime('%d/%m/%Y'); time_str = dt.strftime('%H:%M:%S')
        except:
            date_str = l.get('date','—'); time_str = l.get('heure','—')

        ut = l.get('userType','')
        who = l.get('benevoleNom') or l.get('specNom') or '—'
        mt  = l.get('montant', None)

        vals = [
            date_str, time_str, l.get('action','—'),
            f"{USER_ICON.get(ut,'📝')} {ut}" if ut else '—',
            l.get('label','—'), who,
            l.get('staff','—'), l.get('resaCode','—'),
            euro(mt) if mt is not None else '—',
        ]
        style = S['td_alt'] if i % 2 else S['td']
        for j, val in enumerate(vals):
            cell = ws.cell(row=row, column=j+1, value=val); apply(cell, style)
            if j == 8 and val != '—':
                cell.number_format = '#,##0.00 "€"'
                cell.alignment = Alignment(horizontal='right', vertical='center')
        set_row_height(ws, row, 18)

    ws.freeze_panes = 'A4'
    return ws

# ── MAIN ─────────────────────────────────────────────────────────────
def generate(data_path, output_path):
    with open(data_path) as f:
        data = json.load(f)

    brand = data['event'].get('couleur', '#1a6b7a')
    S = make_styles(brand)

    wb = Workbook()
    sheet_garde(wb, data, S)
    sheet_transactions(wb, data, S)
    sheet_spectateurs(wb, data, S)
    sheet_reservations(wb, data, S)
    sheet_menu(wb, data, S)
    sheet_audit(wb, data, S)

    wb.save(output_path)
    print(f"✅ Rapport généré : {output_path}")

if __name__ == '__main__':
    generate(sys.argv[1], sys.argv[2])
